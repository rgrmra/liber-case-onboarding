import pandas as pd
import numpy
from openpyxl import load_workbook

def read_file(filename):
    df = []
    with open(filename, 'r') as file:
        lines = file.readlines()

        header = lines[0].strip().split(',')

        for line in lines[1:]:
            values = line.strip().split(',')

            if len(values) > len(header):
                values[len(header) - 1] = ','.join(values[len(header) - 1:])
                values = values[:len(header)]
            df.append(values)
    return pd.DataFrame(df, columns=header)

df = read_file("clientes.csv")
df2 = read_file("clientes2.csv")
df = pd.merge(df, df2, on="id") 

#df.rename(columns={"segmento":"sgmento"}, inplace=True)

#xl = pd.read_excel("resultado.xlsx")
#xl = pd.concat([xl, df], axis=0)
#print(xl)
#xl = xl.drop_duplicates(subset=['email'], keep="first")
#print(xl)
#
#xl.to_excel("resultado.xlsx", index=False)

#print(type(5))
#
#df['receita_anual'] = df['receita_anual'].astype(int)
#df['receita_anual'] = df['receita_anual'].map("R$ {:,.2f}".format)

print(df)
res = load_workbook("resultado.xlsx")
A = res['Sheet1']

for index in range(len(df)):
    A.cell(row=index + 2, column=1, value=df['id'][index])
    A.cell(row=index + 2, column=2, value=df['nome'][index])
    A.cell(row=index + 2, column=3, value=df['telefone'][index])
    A.cell(row=index + 2, column=4, value=df['email'][index])
    A.cell(row=index + 2, column=5, value=df['endereco'][index])
    A.cell(row=index + 2, column=6, value=df['cnpj'][index])
    A.cell(row=index + 2, column=7, value=df['segmento'][index])
    A.cell(row=index + 2, column=8, value=df['receita_anual'][index])

res.save("resultado.xlsx")
